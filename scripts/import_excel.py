#!/usr/bin/env python3
"""
DAT Holdings History Excel Importer

Imports historical holdings data from Excel into Supabase holdings_history table.

Usage:
    python scripts/import_excel.py --file path/to/excel.xlsx
    python scripts/import_excel.py --file path/to/excel.xlsx --dry-run
    python scripts/import_excel.py --file path/to/excel.xlsx --company MSTR
"""

import argparse
import re
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

# Add project root to path for imports
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.utils.database import get_db

# Sheet name to ticker mapping
SHEET_MAP = {
    "MSTR": "MSTR",
    "MTPLF": "MTPLF",
    "NAKA": "NAKA",
    "XXI": "XXI",
    "ASST": "ASST",
    "BSTR": "BSTR",
    "ABTC": "ABTC",
    "BMNR": "BMNR",
    "SBET": "SBET",
    "BTBT": "BTBT",
    "BTCS": "BTCS",
    "FGNX": "FGNX",
    "ETHM": "ETHM",
    "ETHZ": "ETHZ",
    "DFDV": "DFDV",
    "HODL": "HODL",
    "UPXI": "UPXI",
    "HSDT": "HSDT",
    "FORD": "FWDI",
    "STSS": "STSS",
    "BREA": "BREA",
    "Hyperion DefI": "HYPD",
    "PURR": "PURR",
    "BNC": "BNC",
}

# Allowed tickers for new company creation (security whitelist)
ALLOWED_NEW_TICKERS = set(SHEET_MAP.values())

# Sheets to skip (patterns)
SKIP_PATTERNS = ["Preferred", "Overview", "->", "v2"]

# Column mapping: Excel header -> database column
COLUMN_MAP = {
    "date": "date",
    "num_of_tokens": "token_count",
    "price": "token_price",
    "nav": "nav",
    "share_price": "share_price",
    "num_of_shares": "shares_outstanding",
}

# MSTR cutoff date
MSTR_CUTOFF = date(2023, 1, 1)


def should_skip_sheet(sheet_name: str) -> bool:
    """Check if sheet should be skipped."""
    for pattern in SKIP_PATTERNS:
        if pattern.lower() in sheet_name.lower():
            return True
    return sheet_name not in SHEET_MAP


def clean_value(value) -> Optional[float]:
    """Clean a cell value, returning None for invalid values."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "" or value == "#REF!" or value == "#N/A" or value == "#VALUE!":
            return None
        try:
            # Remove commas and try to parse
            return float(value.replace(",", ""))
        except (ValueError, AttributeError):
            return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def parse_date(value) -> Optional[date]:
    """Parse a date value from Excel."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # Try common date formats
        for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y"]:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def find_header_row(ws, max_rows: int = 20) -> tuple[int, dict]:
    """
    Find the header row and return (row_index, column_mapping).
    Looks for a row containing 'date' in column A or B.
    """
    for row_idx in range(1, max_rows + 1):
        for col_idx in range(1, 3):  # Check columns A and B
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and str(cell_value).strip().lower() == "date":
                # Found header row, build column mapping
                col_map = {}
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=row_idx, column=col).value
                    if header:
                        header_clean = str(header).strip().lower().replace(" ", "_")
                        if header_clean in COLUMN_MAP:
                            col_map[COLUMN_MAP[header_clean]] = col
                        elif header_clean == "date":
                            col_map["date"] = col
                return row_idx, col_map
    return -1, {}


def import_sheet(
    ws,
    sheet_name: str,
    ticker: str,
    db,
    dry_run: bool = False,
    allow_new_companies: bool = False
) -> dict:
    """
    Import data from a single sheet.
    Returns stats dict with counts.

    Args:
        ws: Worksheet object
        sheet_name: Name of the Excel sheet
        ticker: Company ticker symbol
        db: Database connection
        dry_run: If True, don't write to database
        allow_new_companies: If True, allow creating new companies not in whitelist
    """
    stats = {
        "ticker": ticker,
        "rows_imported": 0,
        "rows_skipped": 0,
        "skip_reasons": {
            "no_date": 0,
            "no_tokens": 0,
            "before_cutoff": 0,
            "invalid_data": 0,
            "duplicate": 0,
        },
        "date_range": {"earliest": None, "latest": None},
        "errors": [],
    }

    # Find header row
    header_row, col_map = find_header_row(ws)
    if header_row < 0 or "date" not in col_map:
        stats["errors"].append(f"Could not find header row with 'date' column")
        return stats

    if "token_count" not in col_map:
        stats["errors"].append(f"Could not find 'num_of_tokens' column")
        return stats

    # Get company_id
    if dry_run:
        company_id = -1  # Placeholder for dry run
    else:
        company_id = db.get_company_id(ticker)
        if not company_id:
            # Validate ticker against whitelist before creating
            if ticker not in ALLOWED_NEW_TICKERS and not allow_new_companies:
                stats["errors"].append(
                    f"Company '{ticker}' not found in database and not in allowed list. "
                    f"Use --allow-new-companies flag to create new companies."
                )
                return stats

            # Validate ticker format (uppercase letters, 2-5 chars)
            if not re.match(r'^[A-Z]{2,5}$', ticker):
                stats["errors"].append(
                    f"Invalid ticker format '{ticker}'. "
                    f"Tickers must be 2-5 uppercase letters (e.g., MSTR, BTC)."
                )
                return stats

            # Create company if it doesn't exist
            try:
                result = db.supabase.table("companies").insert({
                    "ticker": ticker,
                    "name": ticker,  # Will need to update manually
                    "primary_token": "BTC",  # Default, update manually
                }).execute()
                company_id = result.data[0]["id"]
                print(f"    Created new company: {ticker} (id={company_id})")
            except Exception as e:
                stats["errors"].append(f"Failed to create company: {e}")
                return stats

    # Collect all rows first for token_change calculation
    rows_data = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Parse date
        date_val = parse_date(ws.cell(row=row_idx, column=col_map["date"]).value)
        if not date_val:
            stats["rows_skipped"] += 1
            stats["skip_reasons"]["no_date"] += 1
            continue

        # Parse token count
        token_count = clean_value(ws.cell(row=row_idx, column=col_map["token_count"]).value)
        if token_count is None:
            stats["rows_skipped"] += 1
            stats["skip_reasons"]["no_tokens"] += 1
            continue

        # Skip MSTR rows before cutoff
        if ticker == "MSTR" and date_val < MSTR_CUTOFF:
            stats["rows_skipped"] += 1
            stats["skip_reasons"]["before_cutoff"] += 1
            continue

        # Parse optional columns
        row_data = {
            "company_id": company_id,
            "date": date_val,
            "token_count": token_count,
            "source": "excel_import",
        }

        if "token_price" in col_map:
            row_data["token_price"] = clean_value(
                ws.cell(row=row_idx, column=col_map["token_price"]).value
            )

        if "nav" in col_map:
            row_data["nav"] = clean_value(
                ws.cell(row=row_idx, column=col_map["nav"]).value
            )

        if "share_price" in col_map:
            row_data["share_price"] = clean_value(
                ws.cell(row=row_idx, column=col_map["share_price"]).value
            )

        if "shares_outstanding" in col_map:
            row_data["shares_outstanding"] = clean_value(
                ws.cell(row=row_idx, column=col_map["shares_outstanding"]).value
            )

        # Calculate market_cap if both values present
        if row_data.get("share_price") and row_data.get("shares_outstanding"):
            row_data["market_cap"] = row_data["share_price"] * row_data["shares_outstanding"]

        rows_data.append(row_data)

    # Sort by date for token_change calculation
    rows_data.sort(key=lambda x: x["date"])

    # Calculate token_change
    prev_tokens = None
    for row in rows_data:
        if prev_tokens is not None:
            row["token_change"] = row["token_count"] - prev_tokens
        else:
            row["token_change"] = None  # First record, no change
        prev_tokens = row["token_count"]

    # Update date range
    if rows_data:
        stats["date_range"]["earliest"] = rows_data[0]["date"].isoformat()
        stats["date_range"]["latest"] = rows_data[-1]["date"].isoformat()

    # Batch upsert to database
    if not dry_run and rows_data:
        batch_size = 100
        for i in range(0, len(rows_data), batch_size):
            batch = rows_data[i:i + batch_size]

            # Convert dates to strings for JSON
            for row in batch:
                row["date"] = row["date"].isoformat()

            try:
                # Upsert with ON CONFLICT
                db.supabase.table("holdings_history").upsert(
                    batch,
                    on_conflict="company_id,date"
                ).execute()
                stats["rows_imported"] += len(batch)
            except Exception as e:
                stats["errors"].append(f"Batch insert error: {e}")
                # Try individual inserts for debugging
                for row in batch:
                    try:
                        db.supabase.table("holdings_history").upsert(
                            [row],
                            on_conflict="company_id,date"
                        ).execute()
                        stats["rows_imported"] += 1
                    except Exception as e2:
                        stats["rows_skipped"] += 1
                        stats["skip_reasons"]["duplicate"] += 1
    elif dry_run:
        stats["rows_imported"] = len(rows_data)

    return stats


def main():
    parser = argparse.ArgumentParser(
        description="Import DAT holdings history from Excel"
    )
    parser.add_argument(
        "--file", "-f",
        required=True,
        help="Path to Excel file"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and validate without writing to database"
    )
    parser.add_argument(
        "--company", "-c",
        help="Import only one company (ticker) for testing"
    )
    parser.add_argument(
        "--allow-new-companies",
        action="store_true",
        help="Allow creating new companies not in the whitelist (use with caution)"
    )
    args = parser.parse_args()

    # Validate file exists
    excel_path = Path(args.file)
    if not excel_path.exists():
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)

    print("=" * 60)
    print("DAT Holdings History Excel Importer")
    print("=" * 60)
    print(f"File: {excel_path}")
    if args.dry_run:
        print("Mode: DRY RUN (no database writes)")
    if args.company:
        print(f"Filter: {args.company} only")
    if args.allow_new_companies:
        print("WARNING: --allow-new-companies is enabled. New companies may be created.")
    print()

    # Connect to database
    if not args.dry_run:
        db = get_db()
        if not db:
            print("ERROR: Failed to connect to database")
            print("Make sure SUPABASE_URL and SUPABASE_KEY are set in .env")
            sys.exit(1)
    else:
        db = None

    # Load workbook
    print("Loading Excel file...")
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=False)
    except Exception as e:
        print(f"ERROR: Failed to load Excel file: {e}")
        sys.exit(1)

    print(f"Found {len(wb.sheetnames)} sheets")
    print()

    # Process sheets
    all_stats = []
    total_imported = 0
    total_skipped = 0

    for sheet_name in wb.sheetnames:
        # Skip sheets not in map or matching skip patterns
        if should_skip_sheet(sheet_name):
            continue

        ticker = SHEET_MAP.get(sheet_name)
        if not ticker:
            continue

        # Filter by company if specified
        if args.company and ticker != args.company:
            continue

        print(f"Processing: {sheet_name} -> {ticker}")

        ws = wb[sheet_name]
        stats = import_sheet(
            ws, sheet_name, ticker, db,
            dry_run=args.dry_run,
            allow_new_companies=args.allow_new_companies
        )
        all_stats.append(stats)

        # Print results
        print(f"  Imported: {stats['rows_imported']} rows")
        if stats['rows_skipped'] > 0:
            print(f"  Skipped: {stats['rows_skipped']} rows")
            for reason, count in stats['skip_reasons'].items():
                if count > 0:
                    print(f"    - {reason}: {count}")
        if stats['date_range']['earliest']:
            print(f"  Date range: {stats['date_range']['earliest']} to {stats['date_range']['latest']}")
        if stats['errors']:
            for error in stats['errors']:
                print(f"  ERROR: {error}")
        print()

        total_imported += stats['rows_imported']
        total_skipped += stats['rows_skipped']

    # Summary
    print("=" * 60)
    print("IMPORT SUMMARY")
    print("=" * 60)
    print(f"Companies processed: {len(all_stats)}")
    print(f"Total rows imported: {total_imported}")
    print(f"Total rows skipped: {total_skipped}")

    if args.dry_run:
        print()
        print("DRY RUN complete - no data was written to database")

    print()


if __name__ == "__main__":
    main()
