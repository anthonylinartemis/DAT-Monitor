#!/usr/bin/env python3
"""
Export DAT holdings data to a multi-tab Excel file.

Usage:
    python scripts/export_excel.py
    python scripts/export_excel.py --output custom_path.xlsx
    python scripts/export_excel.py --json-only
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add project root to path for imports
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Token colors for header formatting
TOKEN_COLORS = {
    "BTC": "f7931a",
    "ETH": "627eea",
    "SOL": "9945ff",
    "HYPE": "00d395",
    "BNB": "f0b90b",
}

TOKENS = ["BTC", "ETH", "SOL", "HYPE", "BNB"]


def load_from_supabase() -> Optional[dict]:
    """Load data from Supabase database."""
    try:
        from scripts.utils.database import get_db

        db = get_db()
        if not db:
            return None

        # Fetch all companies
        companies_result = db.supabase.table("companies").select("*").execute()
        if not companies_result.data:
            print("No companies found in database")
            return None

        # Fetch all holdings
        holdings_result = db.supabase.table("holdings").select("*").execute()

        # Build company lookup by id
        companies_by_id = {c["id"]: c for c in companies_result.data}

        # Get latest holding per company
        latest_holdings = {}
        for h in holdings_result.data:
            company_id = h["company_id"]
            if company_id not in latest_holdings:
                latest_holdings[company_id] = h
            else:
                if h["filing_date"] > latest_holdings[company_id]["filing_date"]:
                    latest_holdings[company_id] = h

        # Organize companies by token
        companies_by_token = {token: [] for token in TOKENS}
        for company in companies_result.data:
            token = company.get("primary_token")
            if token in companies_by_token:
                latest = latest_holdings.get(company["id"])
                company_data = {
                    "ticker": company["ticker"],
                    "name": company["name"],
                    "cik": company.get("cik", ""),
                    "irUrl": company.get("ir_url", ""),
                    "tokens": latest["token_count"] if latest else 0,
                    "change": latest["token_change"] if latest else 0,
                    "lastUpdate": latest["filing_date"] if latest else "",
                }
                companies_by_token[token].append(company_data)

        # Get all holdings with company info for history
        all_holdings = []
        for h in holdings_result.data:
            company = companies_by_id.get(h["company_id"])
            if company:
                all_holdings.append({
                    "ticker": company["ticker"],
                    "token": company.get("primary_token", ""),
                    "holdings": h["token_count"],
                    "change": h["token_change"],
                    "filing_date": h["filing_date"],
                    "source_url": h.get("source_url", ""),
                })

        return {
            "companies": companies_by_token,
            "history": all_holdings,
            "source": "supabase",
        }

    except Exception as e:
        print(f"Supabase connection failed: {e}")
        return None


def load_from_json() -> Optional[dict]:
    """Load data from data.json backup file."""
    json_path = PROJECT_ROOT / "data.json"

    if not json_path.exists():
        print(f"data.json not found at {json_path}")
        return None

    try:
        with open(json_path, "r") as f:
            data = json.load(f)

        # Transform to consistent format
        companies_by_token = {}
        for token in TOKENS:
            companies_by_token[token] = []
            for company in data.get("companies", {}).get(token, []):
                companies_by_token[token].append({
                    "ticker": company.get("ticker", ""),
                    "name": company.get("name", ""),
                    "cik": company.get("cik", ""),
                    "irUrl": company.get("irUrl", ""),
                    "tokens": company.get("tokens", 0),
                    "change": company.get("change", 0),
                    "lastUpdate": company.get("lastUpdate", ""),
                })

        # Build history from recent changes (limited in JSON)
        history = []
        for change in data.get("recentChanges", []):
            history.append({
                "ticker": change.get("ticker", ""),
                "token": change.get("token", ""),
                "holdings": change.get("tokens", 0),
                "change": change.get("change", 0),
                "filing_date": change.get("date", ""),
                "source_url": "",
            })

        return {
            "companies": companies_by_token,
            "history": history,
            "source": "json",
        }

    except Exception as e:
        print(f"Failed to load data.json: {e}")
        return None


def load_data(json_only: bool = False) -> Optional[dict]:
    """Load data with fallback: Supabase first, then data.json."""
    if not json_only:
        print("Attempting to connect to Supabase...")
        data = load_from_supabase()
        if data:
            print("Successfully loaded data from Supabase")
            return data
        print("Falling back to data.json...")
    else:
        print("Using data.json (--json-only flag)")

    data = load_from_json()
    if data:
        print("Successfully loaded data from data.json")
    return data


def apply_header_style(cell, token: Optional[str] = None):
    """Apply header styling to a cell."""
    cell.font = Font(bold=True, color="FFFFFF" if token else "000000")

    if token and token in TOKEN_COLORS:
        cell.fill = PatternFill(start_color=TOKEN_COLORS[token],
                                end_color=TOKEN_COLORS[token],
                                fill_type="solid")
    else:
        cell.fill = PatternFill(start_color="4472C4",
                                end_color="4472C4",
                                fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        bottom=Side(style="thin", color="000000")
    )


def format_number(value):
    """Format number with commas."""
    if isinstance(value, (int, float)):
        return f"{value:,.0f}"
    return value


def auto_size_columns(ws):
    """Auto-size columns based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_value = str(cell.value) if cell.value else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = max(adjusted_width, 10)


def create_token_sheet(wb: Workbook, token: str, companies: list) -> int:
    """Create a sheet for a specific token."""
    ws = wb.create_sheet(title=token)

    headers = ["Ticker", "Name", "Current Holdings", "Last Change", "Last Updated", "CIK", "IR URL"]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell, token)

    # Sort companies by holdings descending
    sorted_companies = sorted(companies, key=lambda x: x.get("tokens", 0), reverse=True)

    # Write data
    for row, company in enumerate(sorted_companies, 2):
        ws.cell(row=row, column=1, value=company.get("ticker", ""))
        ws.cell(row=row, column=2, value=company.get("name", ""))

        holdings_cell = ws.cell(row=row, column=3, value=company.get("tokens", 0))
        holdings_cell.number_format = "#,##0"

        change_cell = ws.cell(row=row, column=4, value=company.get("change", 0))
        change_cell.number_format = "#,##0"

        ws.cell(row=row, column=5, value=company.get("lastUpdate", ""))
        ws.cell(row=row, column=6, value=company.get("cik", ""))
        ws.cell(row=row, column=7, value=company.get("irUrl", ""))

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-size columns
    auto_size_columns(ws)

    return len(sorted_companies)


def create_all_companies_sheet(wb: Workbook, companies_by_token: dict) -> int:
    """Create the 'All Companies' sheet."""
    ws = wb.create_sheet(title="All Companies")

    headers = ["Ticker", "Name", "Token", "Current Holdings", "Last Change", "Last Updated", "CIK", "IR URL"]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # Collect all companies
    all_companies = []
    for token, companies in companies_by_token.items():
        for company in companies:
            all_companies.append({**company, "token": token})

    # Sort by holdings descending
    sorted_companies = sorted(all_companies, key=lambda x: x.get("tokens", 0), reverse=True)

    # Write data
    for row, company in enumerate(sorted_companies, 2):
        ws.cell(row=row, column=1, value=company.get("ticker", ""))
        ws.cell(row=row, column=2, value=company.get("name", ""))
        ws.cell(row=row, column=3, value=company.get("token", ""))

        holdings_cell = ws.cell(row=row, column=4, value=company.get("tokens", 0))
        holdings_cell.number_format = "#,##0"

        change_cell = ws.cell(row=row, column=5, value=company.get("change", 0))
        change_cell.number_format = "#,##0"

        ws.cell(row=row, column=6, value=company.get("lastUpdate", ""))
        ws.cell(row=row, column=7, value=company.get("cik", ""))
        ws.cell(row=row, column=8, value=company.get("irUrl", ""))

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-size columns
    auto_size_columns(ws)

    return len(sorted_companies)


def create_history_sheet(wb: Workbook, history: list) -> int:
    """Create the 'History' sheet with all holdings records."""
    ws = wb.create_sheet(title="History")

    headers = ["Ticker", "Token", "Holdings", "Change", "Filing Date", "Source URL"]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # Sort by filing_date descending
    sorted_history = sorted(history, key=lambda x: x.get("filing_date", ""), reverse=True)

    # Write data
    for row, record in enumerate(sorted_history, 2):
        ws.cell(row=row, column=1, value=record.get("ticker", ""))
        ws.cell(row=row, column=2, value=record.get("token", ""))

        holdings_cell = ws.cell(row=row, column=3, value=record.get("holdings", 0))
        holdings_cell.number_format = "#,##0"

        change_cell = ws.cell(row=row, column=4, value=record.get("change", 0))
        change_cell.number_format = "#,##0"

        ws.cell(row=row, column=5, value=record.get("filing_date", ""))
        ws.cell(row=row, column=6, value=record.get("source_url", ""))

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-size columns
    auto_size_columns(ws)

    return len(sorted_history)


def export_to_excel(data: dict, output_path: Path) -> dict:
    """Export data to Excel file and return summary stats."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    stats = {"sheets": {}}

    # Create token sheets
    for token in TOKENS:
        companies = data["companies"].get(token, [])
        count = create_token_sheet(wb, token, companies)
        stats["sheets"][token] = count

    # Create All Companies sheet
    all_count = create_all_companies_sheet(wb, data["companies"])
    stats["sheets"]["All Companies"] = all_count

    # Create History sheet
    history_count = create_history_sheet(wb, data.get("history", []))
    stats["sheets"]["History"] = history_count

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Save workbook
    wb.save(output_path)

    # Get file size
    stats["file_size"] = output_path.stat().st_size
    stats["output_path"] = str(output_path)

    return stats


def main():
    parser = argparse.ArgumentParser(
        description="Export DAT holdings data to Excel"
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        help="Custom output path for the Excel file"
    )
    parser.add_argument(
        "--json-only",
        action="store_true",
        help="Skip Supabase and use only data.json"
    )

    args = parser.parse_args()

    # Determine output path
    if args.output:
        output_path = Path(args.output)
    else:
        date_str = datetime.now().strftime("%Y-%m-%d")
        output_path = PROJECT_ROOT / "exports" / f"dat-export-{date_str}.xlsx"

    print("=" * 50)
    print("DAT Holdings Excel Export")
    print("=" * 50)

    # Load data
    data = load_data(json_only=args.json_only)

    if not data:
        print("ERROR: Failed to load data from any source")
        sys.exit(1)

    print(f"Data source: {data['source']}")
    print()

    # Export to Excel
    print("Generating Excel file...")
    stats = export_to_excel(data, output_path)

    # Print summary
    print()
    print("=" * 50)
    print("Export Summary")
    print("=" * 50)
    print()
    print("Row counts per sheet:")
    total_rows = 0
    for sheet_name, count in stats["sheets"].items():
        print(f"  {sheet_name:15} {count:,} rows")
        total_rows += count

    print()
    print(f"Total rows:       {total_rows:,}")

    # Format file size
    file_size = stats["file_size"]
    if file_size >= 1024 * 1024:
        size_str = f"{file_size / (1024 * 1024):.2f} MB"
    elif file_size >= 1024:
        size_str = f"{file_size / 1024:.2f} KB"
    else:
        size_str = f"{file_size} bytes"

    print(f"File size:        {size_str}")
    print(f"Output path:      {stats['output_path']}")
    print()
    print("Export complete!")


if __name__ == "__main__":
    main()
